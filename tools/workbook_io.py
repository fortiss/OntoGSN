# -*- coding: utf-8 -*-
"""One definition of the design workbook's columns and styling.

Used by prov_to_workbook.py to write provenance/Design Documentation.xlsx. The layout
outlives the workbook it was first written for: the hand-maintained
'OntoGSN Design Document.xlsx' was retired once the provenance graph became the source
of truth, and the sheet is now generated from that graph, keeping the same columns so a
reader who knew the old file still recognises this one.

Two columns were renamed when the augmentation and the alignment joined the sheet. Only
the OntoGSN rows rest on the GSN Community Standard; the other two rest on an industrial
process model, on PROV-O and SHACL, and on ARGO. "Item in GSN Community Standard" became
"Item in source" and "Page(s)" became "Page(s) / locator", which is what those records
carry instead of a page.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

LIVE_SHEET = "All rows"
ARCHIVE_SHEET = "Archive"

TTL_COL = "Item in OntoGSN TTL"

CONTENT = ["Item in source", "Page(s) / locator",
           "Item in Natural Language", "Reason(s) for in-/exclusion"]
# graph says which of the three vocabularies a row belongs to. It comes first because
# uid is only unique within a graph: each record numbers its decisions from dd-0001, and
# the pair (graph, uid) is what identifies a row.
# uid is the stable identity (it becomes the provenance IRI); row_key is positional
# and may be renumbered when statements are inserted. match_status is derived, so it
# is reported by check_coverage rather than stored here.
LIVE_COLS = (["graph", "uid", "row_key", "part", "section", "language"] + CONTENT +
             [TTL_COL, "nl_checksum"])
ARCHIVE_COLS = LIVE_COLS + ["archived_because"]

WIDTH = {"graph": 22, "uid": 10, "row_key": 12, "part": 13, "section": 20, "language": 10,
         "Item in source": 46, "Page(s) / locator": 14,
         "Item in Natural Language": 60, "Reason(s) for in-/exclusion": 40,
         "Item in OntoGSN TTL": 56, "nl_checksum": 11, "archived_because": 46,
         "kind": 12, "statement": 110}

HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BODY_FONT = Font(size=9)


def _sheet(wb, title, rows, columns, header_colour):
    ws = wb.create_sheet(title)
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c, "") for c in columns])

    for i, c in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTH.get(c, 20)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=header_colour)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def write(path, live, archived):
    """Two sheets. There was a third, listing axioms with no design decision; it is gone
    because tools/prov_check.py reports the same thing against the live files, and a copy
    of that frozen into a workbook goes stale the moment an axiom is documented."""
    wb = Workbook()
    wb.remove(wb.active)
    _sheet(wb, LIVE_SHEET, live, LIVE_COLS, "1F3864")
    _sheet(wb, ARCHIVE_SHEET, archived, ARCHIVE_COLS, "7F7F7F")
    wb.save(path)
