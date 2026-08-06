# -*- coding: utf-8 -*-
"""Generate the human-readable design documentation from the provenance graph.

    python tools/prov_to_workbook.py                  # write provenance/Design Documentation.xlsx
    python tools/prov_to_workbook.py --verify         # also diff it against the old workbook

The provenance Turtle is the source of truth; this workbook is a view of it, for people
who would rather read a spreadsheet than a graph. Nothing reads it back.

`--verify` compares the generated sheet against 'OntoGSN Design Document.xlsx' cell by
cell. It exists to prove the migration lost nothing before the old file is deleted, and it
is expected to report zero differences.
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from rdflib import Graph, Namespace, RDF
from rdflib.namespace import SKOS

import matching
import workbook_io

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROV_DIR = os.path.join(REPO, "provenance")
TBOX = os.path.join(PROV_DIR, "ontogsn-provenance.ttl")
DATA = os.path.join(PROV_DIR, "ontogsn-provenance-data.ttl")
OUT = os.path.join(PROV_DIR, "Design Documentation.xlsx")

P = Namespace("https://w3id.org/OntoGSN/provenance#")
PROV = Namespace("http://www.w3.org/ns/prov#")

SRC = "Item in GSN Community Standard"
PG = "Page(s)"
RSN = "Reason(s) for in-/exclusion"
NL = "Item in Natural Language"
TTL = "Item in OntoGSN TTL"
NONE = "(none)"


def load():
    """The vocabulary and the record, in one graph - the concept labels live in the
    vocabulary, and rdflib does not follow owl:imports."""
    graph = Graph()
    graph.parse(TBOX, format="turtle")
    graph.parse(DATA, format="turtle")
    return graph


def text(graph, subject, predicate):
    value = graph.value(subject, predicate)
    return "" if value is None else str(value)


def label(graph, node):
    return "" if node is None else str(graph.value(node, SKOS.prefLabel) or "")


def rows_from(graph):
    """-> (live rows, archived rows), each a dict keyed by workbook column."""
    live, archived = [], []
    # a retired decision is typed only as the subclass, and this graph is not reasoned over
    decisions = (set(graph.subjects(RDF.type, P.DesignDecision)) |
                 set(graph.subjects(RDF.type, P.RetiredDecision)))
    for decision in decisions:
        retired = (decision, RDF.type, P.RetiredDecision) in graph
        passage = graph.value(decision, PROV.used)
        rationale = graph.value(decision, P.hasRationale)
        statement = graph.value(decision, PROV.generated)

        no_source = graph.value(decision, P.noSourceRecorded)
        no_reason = graph.value(decision, P.noRationaleRecorded)

        row = {
            "uid": str(decision).rsplit("#dd-", 1)[-1],
            "row_key": text(graph, decision, P.positionKey),
            "part": label(graph, graph.value(decision, P.part)),
            "section": label(graph, graph.value(decision, P.section)),
            "language": label(graph, graph.value(decision, P.formalism)),
            SRC: NONE if no_source else text(graph, passage, P.quotedText),
            PG: NONE if no_source else text(graph, passage, P.pageRef),
            NL: text(graph, decision, P.naturalLanguage),
            RSN: NONE if no_reason else text(graph, rationale, P.rationaleText),
            TTL: text(graph, statement, P.statementText),
            "nl_checksum": text(graph, decision, P.nlWrittenFor),
        }
        row["uid"] = "dd-" + row["uid"]
        if retired:
            row["archived_because"] = text(graph, decision, P.retirementReason)
            archived.append(row)
        else:
            live.append(row)

    # the sheets read in the standard's order, which is what the position key encodes
    live.sort(key=lambda r: r["row_key"])
    archived.sort(key=lambda r: r["row_key"])
    return live, archived


def compare(generated, original):
    """-> list of (row_key, column, generated value, original value)."""
    from openpyxl import load_workbook

    def sheet(path, name):
        book = load_workbook(path, data_only=True)
        if name not in book.sheetnames:
            return {}, []
        work = book[name]
        header = [c.value for c in work[1]]
        out = {}
        for values in work.iter_rows(min_row=2, values_only=True):
            if not any(values):
                continue
            row = {h: ("" if v is None else str(v))
                   for h, v in zip(header, values) if h}
            out[row["uid"]] = row
        return out, header

    problems = []
    for name in (workbook_io.LIVE_SHEET, workbook_io.ARCHIVE_SHEET):
        new, header = sheet(generated, name)
        old, old_header = sheet(original, name)
        if header != old_header:
            problems.append(("-", "HEADER", str(header), str(old_header)))
        for uid in sorted(set(old) | set(new)):
            if uid not in new:
                problems.append((uid, name, "MISSING", "present"))
                continue
            if uid not in old:
                problems.append((uid, name, "present", "MISSING"))
                continue
            for column in old_header:
                if not column:
                    continue
                a, b = new[uid].get(column, ""), old[uid].get(column, "")
                if a != b:
                    problems.append((uid, column, a, b))
    return problems


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=OUT)
    ap.add_argument("--verify", action="store_true",
                    help="diff the result against OntoGSN Design Document.xlsx")
    ap.add_argument("--against", default=workbook_io.WORKBOOK)
    ap.add_argument("--show", type=int, default=20)
    args = ap.parse_args()

    graph = load()
    live, archived = rows_from(graph)
    print(f"{len(live)} live + {len(archived)} retired decisions")

    _, orphans = matching.load_and_match(
        [dict(r, _archived=False) for r in live] +
        [dict(r, _archived=True) for r in archived])

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    workbook_io.write(args.out, live, archived, orphans)
    print(f"wrote {args.out} ({os.path.getsize(args.out):,} bytes)")

    if not args.verify:
        return
    if not os.path.exists(args.against):
        # the workbook this was migrated from has been removed, which was the point
        print(f"\nnothing to verify against: {os.path.basename(args.against)} is gone. "
              f"The provenance graph is now the only record.")
        return
    problems = compare(args.out, args.against)
    if not problems:
        print(f"\nverified against {os.path.basename(args.against)}: "
              f"{len(live) + len(archived)} rows, no differences")
        return
    print(f"\n{len(problems)} differences against {os.path.basename(args.against)}:")
    for uid, column, new, old in problems[:args.show]:
        print(f"  {uid}  {column}")
        print(f"     generated: {new[:110]!r}")
        print(f"     original : {old[:110]!r}")
    if len(problems) > args.show:
        print(f"  ... and {len(problems) - args.show} more")
    sys.exit(1)


if __name__ == "__main__":
    main()
