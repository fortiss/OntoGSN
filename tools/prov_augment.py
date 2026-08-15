# -*- coding: utf-8 -*-
"""Record the rest of what OntoGSN is made of: requirements, competency questions,
stored queries, and the files that are generated from other files.

    python tools/prov_augment.py

Writes provenance/ontogsn-provenance-augmentations.ttl. Unlike the design decisions, all
of this is recoverable from the repository, so this file IS regenerated - re-run it after
adding a query or rebuilding a serialization.

The point is to make the chains explicit and therefore checkable:

    requirement -> competency question -> query -> the ontology terms it names
                                               -> the axioms it rests on or translates
    ontogsn.ttl -> build.py -> ontogsn.rdf, ontogsn.jsonld

The second query chain reaches into the hand-maintained record: a query points at the
StatementRecord of an axiom, and that record carries the design decision, the rationale
and the passage of the standard behind it. So this script reads
ontogsn-provenance-data.ttl as well as the repository - it cannot point at a decision
without knowing which decisions exist.

The query chain is what this caught first: every stored query used to bind gsn: to a
namespace the ontology does not use, which returns nothing at all and says nothing about
it. The queries were rewritten; tools/query_check.py now refuses that and several related
mistakes on every run, and records the result alongside what this script writes.
"""
import argparse
import glob
import hashlib
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

import matching
import prov_ttl
import ttl_model
from prov_ttl import Lit

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "provenance", "ontogsn-provenance-augmentations.ttl")
RECORD = os.path.join(REPO, "provenance", "ontogsn-provenance-data.ttl")
CQ_WORKBOOK = os.path.join(REPO, "provenance", "Competency Questions.xlsx")
QUERY_DIR = os.path.join(REPO, "queries")

PROV_NS = "https://w3id.org/OntoGSN/provenance#"
# the structural keys that declare a term to exist, as opposed to constraining one that
# already does - see statement_index
DECLARING = ("Class", "ObjectProperty", "DatatypeProperty", "AnnotationProperty",
             "Datatype")

ONTOLOGY_NS = "https://w3id.org/OntoGSN/ontology#"
SHAPES_NS = "https://w3id.org/OntoGSN/shapes#"

PREFIX_RE = re.compile(r"PREFIX\s+(\w*):\s*<([^>]+)>", re.IGNORECASE)
VERSION_RE = re.compile(r"owl:versionInfo\s+\"?([0-9][0-9.]*)\"?")


def sha256(path):
    """Content fingerprint, line-ending independent - see matching.file_checksum."""
    return matching.file_checksum(path)


def slug(text):
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def rel(path):
    return os.path.relpath(path, REPO).replace("\\", "/")


def file_block(iri, path, extra=()):
    pairs = [("a", ["gsnprov:File"]),
             ("gsnprov:path", [Lit(rel(path))]),
             ("gsnprov:fileChecksum", [Lit(sha256(path))])]
    pairs.extend(extra)
    return (iri, pairs)


def declared_version(path):
    # The whole file, not a fixed window. owl:versionInfo sits wherever the ontology header
    # happens to put it, and adding one prefix line was enough to push it past a 4000-character
    # read - at which point every slice was reported as stale against an ontology of "unknown"
    # version, which reads as a real finding and is not one.
    with open(path, encoding="utf-8", errors="replace") as fh:
        match = VERSION_RE.search(fh.read())
    return match.group(1) if match else None


def read_sheet(book, name):
    work = book[name]
    header = [c.value for c in work[1]]
    return [{h: ("" if v is None else str(v).strip())
             for h, v in zip(header, values) if h}
            for values in work.iter_rows(min_row=2, values_only=True) if any(values)]


def statement_index():
    """The record, indexed the two ways a query needs to reach into it.

    -> (rule name -> statement, term -> [statements declaring it])

    Both are looked up in the hand-maintained record rather than recomputed from the
    ontology, because the point of the link is to reach a design decision. A statement
    record is what carries one; the axiom in ontogsn.ttl does not.

    The declaring axiom is the one that says a term exists at all - 'Goal a Class',
    'supportedBy a ObjectProperty'. Everything else said about a term constrains
    something already declared, and a query does not rest on those the same way: it
    rests on the term being a thing it can ask about.
    """
    from rdflib import Graph, Namespace

    P = Namespace(PROV_NS)
    graph = Graph()
    graph.parse(RECORD, format="turtle")

    def qname(iri):
        return "gsnprov:" + str(iri)[len(PROV_NS):]

    by_rule = {str(name): qname(st) for st, name in graph.subject_objects(P.ruleName)}

    by_term = {}
    for st, key in graph.subject_objects(P.structuralKey):
        parts = str(key).split("|")
        if len(parts) == 3 and parts[1] == "a" and parts[2] in DECLARING:
            by_term.setdefault(parts[0], []).append(qname(st))
    # two ontology terms can share a rendering, so a term may resolve to more than one
    # record; sorted, because this file has to come out the same way twice
    return by_rule, {term: sorted(set(sts)) for term, sts in by_term.items()}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=OUT)
    args = ap.parse_args()

    graph = ttl_model.load()
    known = {ttl_model.ln(s) for s in graph.subjects()
             if str(s).startswith(ONTOLOGY_NS)}
    blocks = []

    # --- requirements and competency questions ------------------------------------
    book = load_workbook(CQ_WORKBOOK, data_only=True)
    requirements = read_sheet(book, "Initial Requirements")
    questions = read_sheet(book, "Competency Questions")

    blocks.append(prov_ttl.header(
        f"Requirements ({len(requirements)})",
        "What OntoGSN set out to do, before any of it was built."))
    for row in requirements:
        pairs = [("a", ["gsnprov:Requirement"]),
                 ("rdfs:label", [Lit(row["ID"], lang="en")]),
                 ("gsnprov:requirementText", [Lit(row["Requirement"], lang="en")])]
        applies = [p.strip() for p in row.get("refers to", "").split(",") if p.strip()]
        if applies:
            pairs.append(("gsnprov:appliesTo", [Lit(a) for a in applies]))
        source = row.get("source", "").strip()
        if source and source != "-":
            pairs.append(("dc:source", [Lit(source, lang="en")]))
        blocks.append((f"gsnprov:req-{row['ID']}", pairs))

    # the workbook writes some ids with a non-ASCII dash; the IRI uses a plain one
    def question_iri(identifier):
        return "gsnprov:cq-" + re.sub(r"[^A-Za-z0-9]+", "-", identifier).strip("-")

    blocks.append(prov_ttl.header(
        f"Competency questions ({len(questions)})",
        "What the ontology has to be able to answer, and for whom."))
    by_query = {}
    for row in questions:
        identifier = row["ID"]
        iri = question_iri(identifier)
        text = row["Competency Question"]
        # a few rows repeat their own id at the front of the question text
        text = re.sub(r"^" + re.escape(identifier) + r"\s+", "", text)
        blocks.append((iri, [
            ("a", ["gsnprov:CompetencyQuestion"]),
            ("rdfs:label", [Lit(identifier, lang="en")]),
            ("gsnprov:persona", [Lit(row["Role / Persona"], lang="en")]),
            ("gsnprov:questionText", [Lit(text, lang="en")]),
        ]))
        target = row.get("SPARQL Query File", "").strip()
        if target:
            by_query.setdefault(target, []).append(iri)

    # --- stored queries -------------------------------------------------------------
    paths = sorted(glob.glob(os.path.join(QUERY_DIR, "**", "*.rq"), recursive=True))
    by_rule, by_term = statement_index()
    blocks.append(prov_ttl.header(
        f"Stored queries ({len(paths)})",
        "gsnprov:declaresNamespace records what each file binds its prefixes to, which is\n"
        "how a query pointing at the wrong namespace becomes findable.\n\n"
        "gsnprov:queryText holds the query verbatim, so the record says what was asked at\n"
        "a given version without the file having to still exist. gsnprov:fileChecksum is\n"
        "what detects that it changed.\n\n"
        "Two links reach back into the ontology, and they are not the same strength.\n"
        "gsnprov:translatesStatement is an exact correspondence: the files under\n"
        "queries/rules/ ARE the SWRL rules, re-expressed for a store with no reasoner, and\n"
        "the join is the rule's name. gsnprov:restsOnStatement is weaker and derived: the\n"
        "axiom declaring each gsn: term the query names. Either way the target is a\n"
        "StatementRecord, so one hop further reaches why that axiom exists at all."))
    unanswered = dict(by_query)
    unjoined = []
    for path in paths:
        name = os.path.relpath(path, QUERY_DIR).replace(os.sep, "/")
        with open(path, encoding="utf-8", errors="replace") as fh:
            body = fh.read()
        namespaces = sorted({uri for _, uri in PREFIX_RE.findall(body)})
        mentions = sorted({f"gsn:{local}" for local in
                           re.findall(r"\bgsn:([A-Za-z][A-Za-z0-9_]*)", body)
                           if local in known})
        pairs = [("a", ["gsnprov:Query"]),
                 ("rdfs:label", [Lit(name, lang="en")]),
                 ("gsnprov:formalism", ["gsnprov:SPARQL"]),
                 ("gsnprov:path", [Lit(rel(path))]),
                 ("gsnprov:fileChecksum", [Lit(sha256(path))])]
        if namespaces:
            pairs.append(("gsnprov:declaresNamespace", [Lit(n) for n in namespaces]))
        if mentions:
            pairs.append(("gsnprov:mentionsTerm", mentions))
        rule = re.search(r"(?m)^#\s+name:\s+(.*)$", body)
        if name.startswith("rules/") and rule:
            rule_name = rule.group(1).strip()
            pairs.append(("gsnprov:ruleName", [Lit(rule_name)]))
            # the rule this file carries out, as the ontology states it. Missing only if
            # the rule was renamed on one side and not the other, which is worth seeing.
            if rule_name in by_rule:
                pairs.append(("gsnprov:translatesStatement", [by_rule[rule_name]]))
            else:
                unjoined.append((name, rule_name))
        rests = sorted({st for term in mentions
                        for st in by_term.get(term.split(":", 1)[1], [])})
        if rests:
            pairs.append(("gsnprov:restsOnStatement", rests))
        for question in by_query.get(name, []):
            pairs.append(("gsnprov:answersQuestion", [question]))
        pairs.append(("gsnprov:queryText", [Lit(body)]))
        unanswered.pop(name, None)
        blocks.append((f"gsnprov:query-{slug(name[:-3])}", pairs))

    # --- generated files -------------------------------------------------------------
    blocks.append(prov_ttl.header(
        "Generated files and what generates them",
        "A derived file with no generating activity is one nobody knows how to rebuild.\n"
        "The slices under serializations/separated/ are exactly that, which is why they\n"
        "appear here with prov:wasDerivedFrom but no prov:wasGeneratedBy."))

    source_ttl = os.path.join(REPO, "serializations", "ontogsn.ttl")
    blocks.append(file_block("gsnprov:file-ontogsn-ttl", source_ttl, [
        ("gsnprov:declaredVersion", [Lit(declared_version(source_ttl) or "unknown")]),
        ("prov:specializationOf", ["<https://w3id.org/OntoGSN/ontology>"]),
    ]))

    derived = []
    for name in ("ontogsn.rdf", "ontogsn.jsonld"):
        path = os.path.join(REPO, "serializations", name)
        iri = "gsnprov:file-" + slug(name)
        derived.append(iri)
        blocks.append(file_block(iri, path, [
            ("prov:wasDerivedFrom", ["gsnprov:file-ontogsn-ttl"]),
            ("prov:specializationOf", ["<https://w3id.org/OntoGSN/ontology>"]),
            ("prov:wasGeneratedBy", ["gsnprov:build-serializations"]),
        ]))
    blocks.append(("gsnprov:build-serializations", [
        ("a", ["prov:Activity"]),
        ("rdfs:label", [Lit("serializations/build.py", lang="en")]),
        ("prov:used", ["gsnprov:file-ontogsn-ttl"]),
        ("prov:generated", derived),
        ("prov:wasAssociatedWith", ["gsnprov:agent-build-py"]),
    ]))
    blocks.append(("gsnprov:agent-build-py", [
        ("a", ["prov:SoftwareAgent", "prov:Plan"]),
        ("rdfs:label", [Lit("serializations/build.py", lang="en")]),
        ("gsnprov:path", [Lit("serializations/build.py")]),
    ]))

    sections = sorted(glob.glob(os.path.join(REPO, "shapes",
                                             "ontogsn-shapes_[1-5]*.ttl")))
    section_iris = []
    for path in sections:
        iri = "gsnprov:file-" + slug(os.path.basename(path)[:-4])
        section_iris.append(iri)
        blocks.append(file_block(iri, path, [
            ("gsnprov:declaredVersion", [Lit(declared_version(path) or "unknown")]),
            ("prov:wasDerivedFrom", ["gsnprov:file-ontogsn-ttl"]),
        ]))
    full = os.path.join(REPO, "shapes", "ontogsn-shapes_0_full.ttl")
    blocks.append(file_block("gsnprov:file-ontogsn-shapes-0-full", full, [
        ("gsnprov:declaredVersion", [Lit(declared_version(full) or "unknown")]),
        ("prov:wasDerivedFrom", section_iris),
        ("prov:wasGeneratedBy", ["gsnprov:build-shapes"]),
    ]))
    blocks.append(("gsnprov:build-shapes", [
        ("a", ["prov:Activity"]),
        ("rdfs:label", [Lit("shapes/build_full.py", lang="en")]),
        ("prov:used", section_iris),
        ("prov:generated", ["gsnprov:file-ontogsn-shapes-0-full"]),
        ("prov:wasAssociatedWith", ["gsnprov:agent-build-full-py"]),
    ]))
    blocks.append(("gsnprov:agent-build-full-py", [
        ("a", ["prov:SoftwareAgent", "prov:Plan"]),
        ("rdfs:label", [Lit("shapes/build_full.py", lang="en")]),
        ("gsnprov:path", [Lit("shapes/build_full.py")]),
    ]))

    # The slices are the ontology itself, cut into sections, so they are specializations
    # of it and owe it their version number. The shapes above are a different graph with
    # a version of their own, which is why they are derived from the ontology but not
    # specializations of it.
    slices = sorted(glob.glob(os.path.join(REPO, "serializations", "separated",
                                           "*.ttl")) +
                    glob.glob(os.path.join(REPO, "serializations", "separated",
                                           "*.jsonld")))
    slice_iris = []
    for path in slices:
        version = declared_version(path)
        iri = "gsnprov:file-" + slug(os.path.basename(path).rsplit(".", 1)[0] + "-" +
                                     path.rsplit(".", 1)[1])
        slice_iris.append(iri)
        extra = [("prov:wasDerivedFrom", ["gsnprov:file-ontogsn-ttl"]),
                 ("prov:specializationOf", ["<https://w3id.org/OntoGSN/ontology>"]),
                 ("prov:wasGeneratedBy", ["gsnprov:build-separated"])]
        if version:
            extra.append(("gsnprov:declaredVersion", [Lit(version)]))
        blocks.append(file_block(iri, path, extra))
    blocks.append(("gsnprov:build-separated", [
        ("a", ["prov:Activity"]),
        ("rdfs:label", [Lit("serializations/build_separated.py", lang="en")]),
        ("prov:used", ["gsnprov:file-ontogsn-ttl"]),
        ("prov:generated", slice_iris),
        ("prov:wasAssociatedWith", ["gsnprov:agent-build-separated-py"]),
    ]))
    blocks.append(("gsnprov:agent-build-separated-py", [
        ("a", ["prov:SoftwareAgent", "prov:Plan"]),
        ("rdfs:label", [Lit("serializations/build_separated.py", lang="en")]),
        ("gsnprov:path", [Lit("serializations/build_separated.py")]),
    ]))

    preamble = (
        "# GENERATED by tools/prov_augment.py. Everything here is recoverable from the\n"
        "# repository, so unlike ontogsn-provenance-data.ttl this file is rebuilt, not\n"
        "# edited. Re-run after adding a query or rebuilding a serialization.\n\n"
        "<https://w3id.org/OntoGSN/provenance/augmentations> a owl:Ontology ;\n"
        "    owl:imports <https://w3id.org/OntoGSN/provenance> ;\n"
        "    dc:title \"OntoGSN Provenance - requirements, questions, queries and builds\"@en ;\n"
        "    owl:versionInfo \"1.0.0\" .\n\n")

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    prov_ttl.write(args.out, blocks, preamble)

    print(f"{len(requirements)} requirements, {len(questions)} competency questions, "
          f"{len(paths)} queries")
    print(f"{len(sections) + len(slices) + 4} files with their derivation chains")
    if unanswered:
        print(f"\n{len(unanswered)} competency questions name a query file that does "
              f"not exist:")
        for name, who in sorted(unanswered.items()):
            print(f"  {name}  (cited by {', '.join(w.split(':')[-1] for w in who)})")
    if unjoined:
        # a rule with no record is either a rule nobody documented or a rename that
        # happened on one side only. Both need a person, so say so rather than dropping
        # the link silently.
        print(f"\n{len(unjoined)} rule queries name a rule with no statement record:")
        for name, rule_name in sorted(unjoined):
            print(f"  {name}  ->  {rule_name!r}")
    print(f"\nwrote {args.out} ({os.path.getsize(args.out):,} bytes)")


if __name__ == "__main__":
    main()
